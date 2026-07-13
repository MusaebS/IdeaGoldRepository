"""Report builders: fairness/audit frames and the Excel / PDF exports.

The downloadable report is treated as a first-class deliverable: a titled,
legended, print-friendly PDF (landscape A4) and a styled Excel workbook that
carry the same information as the screen — schedule grid, per-role fairness,
annotations — without the raw-dump look. Pure "print view" helpers shape the
data (testable without reportlab/openpyxl); the two byte-builders only lay
that shaped data out.
"""
from __future__ import annotations

import io
import json
import math
import os
import unicodedata
from datetime import date as _date
from pathlib import Path
from typing import Dict, List, Mapping, Sequence, Tuple
from xml.sax.saxutils import escape

try:
    import pandas as pd
except ImportError:  # pragma: no cover - fallback when pandas missing
    from .pandas_stub import pd

from .coloring import DEFAULT_PALETTE, schedule_cell_colors
from .data_models import InputData
from .fairness import ResidentPoints, calculate_points
from .points import classify_slot
from .utils import compact_date_range, friendly_date, weekend_holiday_dates

__all__ = [
    "build_fairness_frame",
    "build_assignment_frame",
    "build_cumulative_frame",
    "build_policy_snapshot_frame",
    "spreadsheet_safe_text",
    "spreadsheet_safe_frame",
    "schedule_to_excel_bytes",
    "schedule_to_pdf_bytes",
]


def _is_missing(value) -> bool:
    if value is None:
        return True
    if isinstance(value, float):
        return math.isnan(value)
    checker = getattr(pd, "isna", None)
    if checker is not None:
        try:
            return bool(checker(value))
        except (TypeError, ValueError):
            pass
    return False


def _fmt(value) -> str:
    if _is_missing(value):
        return ""
    if isinstance(value, float):
        return f"{value:g}"
    return str(value)


def spreadsheet_safe_text(value):
    """Return a cell/CSV-safe scalar without changing typed numbers or dates.

    Spreadsheet applications can execute text beginning with ``=``, ``+``,
    ``-``, ``@``, a tab, or a carriage return as a formula. Prefixing an
    apostrophe makes the value literal; Excel hides that prefix when it
    displays the cell. Leading whitespace is also inspected because it can be
    ignored by importers before formula evaluation.
    """
    if not isinstance(value, str) or not value:
        return value
    stripped = value.lstrip(" \t\r\n")
    if value[0] in "\t\r\n" or (stripped and stripped[0] in "=+-@"):
        return "'" + value
    return value


def spreadsheet_safe_frame(frame):
    """Copy a DataFrame and neutralise formula-like values and headers."""
    safe = frame.copy()
    safe.columns = [spreadsheet_safe_text(column) for column in safe.columns]
    for index in range(len(safe.columns)):
        safe.isetitem(
            index,
            [spreadsheet_safe_text(value) for value in safe.iloc[:, index]],
        )
    return safe


_PDF_PUNCTUATION = str.maketrans({
    "–": "-", "—": "-", "‑": "-", "→": "->", "×": "x", "·": "|",
    "“": '"', "”": '"', "‘": "'", "’": "'", "…": "...",
})


def _pdf_safe_text(value, *, unicode_font: bool = False) -> str:
    """Render text with a Unicode font, or a reversible ASCII fallback.

    When a suitable installed font is available, Arabic text is reshaped and
    bidi-ordered before ReportLab sees it. If the host has no Unicode font,
    unsupported characters are preserved as reversible ``U+XXXX`` tokens
    rather than disappearing. Excel always retains the original Unicode text.
    """
    text = _fmt(value).translate(_PDF_PUNCTUATION)
    if unicode_font:
        # ReportLab does not shape right-to-left scripts itself. The small
        # optional runtime helpers are direct dependencies of the app, but the
        # fallback keeps local/model-only environments functional.
        try:
            import arabic_reshaper
            from bidi.algorithm import get_display

            if any("\u0600" <= char <= "\u06ff" for char in text):
                text = get_display(arabic_reshaper.reshape(text))
        except ImportError:
            pass
        return text
    out: List[str] = []
    for char in unicodedata.normalize("NFKD", text):
        if ord(char) < 128:
            out.append(char)
        elif unicodedata.combining(char):
            continue
        else:
            out.append(f"[U+{ord(char):04X}]")
    return "".join(out)


def _register_pdf_fonts():
    """Register an installed Unicode font, with safe base-font fallback."""
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    windows_fonts = Path(os.environ.get("WINDIR", "C:/Windows")) / "Fonts"
    candidates = [
        (
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
            Path("/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
        ),
        (
            windows_fonts / "arial.ttf",
            windows_fonts / "arialbd.ttf",
        ),
        (
            Path("/System/Library/Fonts/Supplemental/Arial.ttf"),
            Path("/System/Library/Fonts/Supplemental/Arial Bold.ttf"),
        ),
    ]
    for normal_path, bold_path in candidates:
        if not normal_path.is_file():
            continue
        try:
            pdfmetrics.registerFont(TTFont("IdeaGoldUnicode", str(normal_path)))
            if bold_path.is_file():
                pdfmetrics.registerFont(TTFont("IdeaGoldUnicodeBold", str(bold_path)))
                bold_name = "IdeaGoldUnicodeBold"
            else:
                bold_name = "IdeaGoldUnicode"
        except Exception:
            continue
        return "IdeaGoldUnicode", bold_name, True
    return "Helvetica", "Helvetica-Bold", False


def build_fairness_frame(
    points: Dict[str, ResidentPoints],
    data: InputData,
    df=None,
    prior_ledger=None,
    ledger_policy=None,
) -> "pd.DataFrame":
    """Return a per-resident fairness table (total, weekend, NF, per-label).

    When the solved ``df`` is given, target and deviation columns are added
    from the same solver-resolved targets the fairness log uses (one source of
    truth), plus per-label call *counts* (``"<label> n"``). ``prior_ledger``
    adds ``Prior …`` / ``Cumulative …`` columns (and cumulative call counts
    when the ledger carries a per-label history), showing the multi-block
    picture the carryover balancing works from. A ``Notes`` column carries the
    same load annotations as the fairness log (groups, perks, exemptions,
    blackouts, reductions, leaves).
    """
    from .fairness import (  # shared target resolution / annotations
        _resolved_target,
        calculate_label_counts,
        load_annotation_notes,
        preference_satisfaction,
    )

    target_total = _resolved_target(df, "target_total", data.target_total) if df is not None else None
    target_total_map = _resolved_target(df, "target_total_map", data.target_total_map) if df is not None else None
    target_weekend = _resolved_target(df, "target_weekend", data.target_weekend) if df is not None else None
    target_label = _resolved_target(df, "target_label", data.target_label) if df is not None else data.target_label
    counts = calculate_label_counts(df, data) if df is not None else None
    pref_stats = preference_satisfaction(df, data) if df is not None else {}

    labels = sorted(
        {shift.label for shift in data.shifts}
        | {label for info in points.values() for label in info.get("labels", {})}
        | {label for (_person, label) in (target_label or {})}
    )
    prior = prior_ledger or {}
    prior_has_counts = any((entry or {}).get("label_counts") for entry in prior.values())
    ending_ledger = None
    show_cumulative = bool(prior)
    if df is not None:
        from .ledger import update_ledger

        ending_ledger = update_ledger(prior, df, data, policy=ledger_policy)
        show_cumulative = show_cumulative or any(
            abs(
                float((ending_ledger.get(name) or {}).get(dim, 0.0))
                - (
                    float((prior.get(name) or {}).get(dim, 0.0))
                    + (
                        float(info.get("total", 0.0))
                        if dim == "total"
                        else float(info.get("weekend", 0.0))
                    )
                )
            ) > 1e-9
            for name, info in points.items()
            for dim in ("total", "weekend")
        )
    rows = []
    for name in sorted(points):
        info = points[name]
        role = "Senior" if name in data.seniors else "Junior"
        row = {
            "Resident": name,
            "Role": role,
            # Points (weighted: a weekend / heavier shift is worth more than 1).
            "Total points": info.get("total", 0.0),
            "Weekend points": info.get("weekend", 0.0),
            # Plain shift counts (how many calls), alongside the points so the
            # two are never confused: e.g. 7 points can be 7 day shifts or a
            # smaller number of heavier weekend shifts.
            "Shifts": int(info.get("total_calls", 0)),
            "Weekend shifts": int(info.get("weekend_calls", 0)),
            # NF duty is a day count (coverage overlay), outside regular fairness.
            "NF duty (days)": int(info.get("night_float", 0.0)),
        }
        total_tgt = (target_total_map or {}).get(name, target_total)
        if total_tgt is not None:
            row["Total target"] = round(total_tgt, 1)
            row["Total dev"] = round(info.get("total", 0.0) - total_tgt, 1)
        if target_weekend and name in target_weekend:
            row["Weekend target"] = round(target_weekend[name], 1)
            row["Weekend dev"] = round(info.get("weekend", 0.0) - target_weekend[name], 1)
        for label in labels:
            label_actual = float(info.get("labels", {}).get(label, 0.0))
            row[label] = label_actual
            if counts is not None:
                row[f"{label} n"] = counts.get(name, {}).get(label, 0)
            if target_label and (name, label) in target_label:
                label_target = float(target_label[(name, label)])
                row[f"{label} target"] = round(label_target, 1)
                row[f"{label} dev"] = round(label_actual - label_target, 1)
        if show_cumulative:
            prior_entry = prior.get(name) or {}
            current = {
                "total": info.get("total", 0.0),
                "weekend": info.get("weekend", 0.0),
            }
            for col, dim in (("total", "total"), ("weekend", "weekend")):
                before = float(prior_entry.get(dim, 0.0))
                row[f"Prior {col}"] = round(before, 1)
                row[f"Cumulative {col}"] = round(before + current[dim], 1)
                if ending_ledger is not None:
                    row[f"Policy-adjusted cumulative {col}"] = round(
                        float((ending_ledger.get(name) or {}).get(dim, 0.0)), 1
                    )
            row["Cumulative basis"] = (
                "Cumulative = prior ledger + actual points; policy-adjusted "
                "columns are the standing saved for the next block"
            )
            if prior_has_counts and counts is not None:
                prior_counts = prior_entry.get("label_counts") or {}
                for label in labels:
                    row[f"{label} n cum"] = (
                        int(prior_counts.get(label, 0)) + counts.get(name, {}).get(label, 0)
                    )
        if name in pref_stats:
            matched, total = pref_stats[name]
            row["Pref match"] = f"{matched}/{total}"
        notes = load_annotation_notes(name, data)
        if notes:
            row["Notes"] = " ".join(notes)
        rows.append(row)
    return pd.DataFrame(rows)


def build_assignment_frame(df, data: InputData) -> "pd.DataFrame":
    """One row per (date, shift) slot: the per-call audit detail.

    ``Points`` is the regular workload actually awarded. ``Nominal points``
    records what the slot would have been worth as regular demand. Closed,
    unfilled, and night-float-overlay cells award zero regular points. An
    NF-eligible shift without an active overlay remains an ordinary regular
    assignment and awards its nominal points.
    """
    from .closures import closed_cells_from_attr
    from .night_float import nf_cells_from_attr

    weekend_dates = weekend_holiday_dates(data)
    closed_cells = closed_cells_from_attr(df)
    nf_cells = nf_cells_from_attr(df)
    rows = []
    for record in df.to_dict("records"):
        day = record.get("Date")
        day_key = day.isoformat() if hasattr(day, "isoformat") else day
        for sh in data.shifts:
            slot = classify_slot(day, sh, data, weekend_dates)
            person = record.get(sh.label)
            key = (day_key, sh.label)
            is_closed = key in closed_cells or person == "Closed"
            is_overlay = key in nf_cells
            missing = person in (None, "Unfilled") or not isinstance(person, str)
            if is_closed:
                status = "Closed"
                resident = "Closed"
                awarded = 0.0
            elif is_overlay:
                resident = "Unfilled" if missing else person
                expected = nf_cells.get(key)
                if missing:
                    status = "Invalid NF overlay (unfilled)"
                elif expected and person != expected:
                    status = "Invalid NF overlay assignment"
                else:
                    status = "Night float overlay"
                awarded = 0.0
            elif missing:
                status = "Unfilled"
                resident = "Unfilled"
                awarded = 0.0
            else:
                status = "Regular assignment"
                resident = person
                awarded = slot.points
            rows.append({
                "Date": day,
                "Day": record.get("Day") or (day.strftime("%A") if hasattr(day, "strftime") else ""),
                "Shift": sh.label,
                "Status": status,
                "Resident": resident,
                "Points": awarded,
                "Nominal points": slot.points,
                "Weekend": slot.weekend,
                "Night float": is_overlay,
                "NF-eligible shift": slot.night_float,
            })
    return pd.DataFrame(rows)


def build_cumulative_frame(
    points: Dict[str, ResidentPoints],
    prior_ledger,
    data: InputData,
    ledger_policy=None,
) -> "pd.DataFrame":
    """Long-form rows for the cumulative standing chart.

    One row per resident per segment (``Prior blocks`` from the uploaded
    ledger, ``This block`` from the solved schedule) with the resident's role
    and cumulative total, so a stacked bar chart can show how this block
    builds on history. Residents that appear only in the ledger (departed)
    are skipped — the chart describes the people being scheduled now.
    """
    from .ledger import DEFAULT_POLICY, block_adjustments

    prior = prior_ledger or {}
    policy = DEFAULT_POLICY if ledger_policy is None else ledger_policy
    adjustments = block_adjustments(prior, data)
    rows = []
    for name in sorted(points):
        before = float((prior.get(name) or {}).get("total", 0.0))
        current = float(points[name].get("total", 0.0))
        role = "Senior" if name in data.seniors else "Junior"
        cumulative = round(before + current, 1)
        adjustment = adjustments.get(name) or {}
        policy_delta = 0.0
        if policy.no_refund_penalties:
            policy_delta -= float(adjustment.get("penalty", 0.0))
        if policy.no_catchup_excused:
            policy_delta += float(adjustment.get("excused_total", 0.0))
        ledger_standing = round(before + current + policy_delta, 1)
        rows.append({
            "Resident": name, "Role": role, "Segment": "Prior blocks",
            "Points": round(before, 1), "Cumulative": cumulative,
            "Policy adjustment": round(policy_delta, 1),
            "Policy-adjusted cumulative": ledger_standing,
            "Standing basis": "Actual points before next-block ledger policy",
        })
        rows.append({
            "Resident": name, "Role": role, "Segment": "This block",
            "Points": round(current, 1), "Cumulative": cumulative,
            "Policy adjustment": round(policy_delta, 1),
            "Policy-adjusted cumulative": ledger_standing,
            "Standing basis": "Actual points before next-block ledger policy",
        })
    return pd.DataFrame(rows)


def build_policy_snapshot_frame(
    data: InputData,
    df=None,
    validation_issues: Sequence[str] | None = None,
    extra: Mapping[str, object] | None = None,
    ledger_policy=None,
    *,
    include_config_details: bool = False,
    prior_ledger=None,
) -> "pd.DataFrame":
    """Return a compact, auditable snapshot of the policy behind an export."""
    from .ledger import DEFAULT_POLICY

    attrs = getattr(df, "attrs", {}) or {}
    active_ledger_policy = DEFAULT_POLICY if ledger_policy is None else ledger_policy
    weekend_days = data.weekend_days if data.weekend_days is not None else [5, 6]
    weekday_names = ("Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday")

    def _weekday_label(day) -> str:
        try:
            index = int(day)
        except (TypeError, ValueError):
            return f"Invalid ({day})"
        return weekday_names[index] if 0 <= index < len(weekday_names) else f"Invalid ({day})"

    shifts = "; ".join(
        f"{s.label} ({s.role}, {s.points:g} base pts"
        + (", NF-eligible" if s.night_float else "")
        + (", Thu weekend" if s.thu_weekend else "")
        + ")"
        for s in data.shifts
    )
    rows = [
        {"Setting": "Schedule period", "Value": f"{data.start_date} to {data.end_date}"},
        {"Setting": "Roster", "Value": f"{len(data.juniors)} juniors; {len(data.seniors)} seniors"},
        {"Setting": "Minimum gap", "Value": f"{data.min_gap} day(s)"},
        {
            "Setting": "Weekend days",
            "Value": ", ".join(_weekday_label(day) for day in weekend_days),
        },
        {"Setting": "Weekend multiplier", "Value": f"x{data.weekend_multiplier:g}"},
        {
            "Setting": "Weekend fairness guardrail",
            "Value": "Target-relative residual spread within each role",
        },
        {"Setting": "Shift definitions", "Value": shifts},
        {"Setting": "Night-float rest", "Value": f"{data.nf_rest_days} day(s)"},
        {"Setting": "Random seed", "Value": data.seed},
        {
            "Setting": "Per-shift-type ledger carryover",
            "Value": (
                "Enabled"
                if attrs.get("label_carryover", True)
                else "Disabled"
            ),
        },
        {
            "Setting": "Solver time limit",
            "Value": (
                f"{attrs['time_limit_sec']:g} seconds"
                if attrs.get("time_limit_sec") is not None
                else "Not recorded"
            ),
        },
        {
            "Setting": "Coverage policy counts",
            "Value": (
                f"{len(data.leaves or [])} leave(s); "
                f"{len(data.rotators or [])} rotator window(s); "
                f"{len(data.nf_assignments or [])} NF assignment(s); "
                f"{len(data.closures or [])} closure(s)"
            ),
        },
        {
            "Setting": "Load policy counts",
            "Value": (
                f"{len(data.max_total or {})} total cap(s); "
                f"{len(data.extra_points or {})} mandatory extra(s); "
                f"{len(data.perks or [])} perk(s); "
                f"{len(data.reductions or [])} reduction(s)"
            ),
        },
        {
            "Setting": "Restriction counts",
            "Value": (
                f"{len(data.blackouts or [])} blackout(s); "
                f"{len(data.exempt_shifts or {})} resident exemption set(s); "
                f"{len(data.avoid_pairs or [])} avoid pair(s)"
            ),
        },
        {
            "Setting": "Ledger: refund penalties later",
            "Value": "No" if active_ledger_policy.no_refund_penalties else "Yes",
        },
        {
            "Setting": "Ledger: catch up excused load later",
            "Value": "No" if active_ledger_policy.no_catchup_excused else "Yes",
        },
        {"Setting": "Solver status", "Value": attrs.get("solver_status") or "Not recorded"},
        {"Setting": "Manually edited", "Value": "Yes" if attrs.get("manually_edited") else "No"},
    ]
    if validation_issues is not None:
        rows.append({
            "Setting": "Validation",
            "Value": "Passed" if not validation_issues else f"{len(validation_issues)} issue(s)",
        })
        rows.extend(
            {"Setting": f"Validation issue {index}", "Value": issue}
            for index, issue in enumerate(validation_issues, start=1)
        )
    for key, value in (extra or {}).items():
        rows.append({"Setting": str(key), "Value": value})
    if include_config_details:
        from .config_io import input_data_to_json

        config_payload = json.loads(input_data_to_json(data))
        rows.extend(
            {
                "Setting": f"Config JSON: {key}",
                "Value": json.dumps(value, ensure_ascii=False, separators=(",", ":")),
            }
            for key, value in config_payload.items()
        )
        rows.extend(
            {
                "Setting": f"Prior ledger JSON: {name}",
                "Value": json.dumps(
                    value, ensure_ascii=False, sort_keys=True, separators=(",", ":")
                ),
            }
            for name, value in sorted((prior_ledger or {}).items())
        )
    return pd.DataFrame(rows)


# --- print views (pure data shaping, shared by the PDF and tested directly) --

def schedule_print_view(df, data: InputData) -> Tuple[List[str], List[dict], set]:
    """Shape the schedule frame for print: columns, rows, weekend row indexes.

    * ``Date`` and ``Day`` merge into one friendly ``Date`` column
      ("Sat 07 Jan") — the two duplicated each other and wasted width.
    * Empty cells in shift columns become an explicit ``Unfilled`` so a
      printed gap can't be mistaken for a formatting accident.
    * Weekend rows (by date, including weekend-flagged holidays) are returned
      for row shading. Row indexes match the input frame, so the per-cell
      colour map stays valid.
    """
    columns = list(df.columns)
    shift_labels = {s.label for s in data.shifts}
    merged = "Date" in columns
    out_columns = (
        ["Date"] + [c for c in columns if c not in ("Date", "Day")]
        if merged
        else columns
    )
    weekend_days = (
        set(data.weekend_days) if data.weekend_days is not None else {5, 6}
    )
    holiday_weekends = weekend_holiday_dates(data)
    weekend_rows: set = set()
    out_rows: List[dict] = []
    for idx, record in enumerate(df.to_dict("records")):
        day = record.get("Date")
        row: dict = {}
        if merged:
            row["Date"] = friendly_date(day)
        for column in out_columns:
            if column == "Date" and merged:
                continue
            value = record.get(column)
            if column in shift_labels and _is_missing(value):
                value = "Unfilled"
            row[column] = value
        out_rows.append(row)
        if hasattr(day, "weekday") and (
            day.weekday() in weekend_days or day in holiday_weekends
        ):
            weekend_rows.add(idx)
    return out_columns, out_rows, weekend_rows


# Print column order for the per-role fairness tables. Raw per-label POINT
# columns and the Notes column are deliberately left to the CSV/Excel: counts
# read better on paper, and notes become numbered footnotes below the table.
_PRINT_LEAD_COLS = (
    "Resident", "Total points", "Total target", "Total dev",
    "Weekend points", "Weekend target", "Weekend dev",
    "Shifts", "Weekend shifts",
    "Prior total", "Cumulative total",
)
_PRINT_TAIL_COLS = ("NF duty (days)", "Pref match")


def fairness_print_sections(
    fairness: "pd.DataFrame", data: InputData
) -> List[Tuple[str, List[str], List[dict]]]:
    """Split the fairness frame into per-role print sections.

    Returns ``[(title, columns, rows)]`` — one section per role present, in
    Junior → Senior order, each with a curated print column set. Columns that
    would be entirely empty/zero (NF duty on a roster without night float,
    Pref match when nobody has preferences) are dropped.
    """
    records = fairness.to_dict("records")
    if not records:
        return []
    shift_roles = {shift.label: shift.role for shift in data.shifts}
    sections: List[Tuple[str, List[str], List[dict]]] = []
    for role, title in (("Junior", "Juniors"), ("Senior", "Seniors")):
        rows = [r for r in records if r.get("Role") == role]
        if not rows:
            continue
        columns = [c for c in _PRINT_LEAD_COLS if c in fairness.columns]
        for label in sorted(lbl for lbl, shift_role in shift_roles.items() if shift_role == role):
            for column in (f"{label} n", f"{label} target", f"{label} dev"):
                if column in fairness.columns and (
                    column.endswith(" n") or any(not _is_missing(r.get(column)) for r in rows)
                ):
                    columns.append(column)
        for col in _PRINT_TAIL_COLS:
            if col in fairness.columns and any(_truthy(r.get(col)) for r in rows):
                columns.append(col)
        sections.append(
            (title, columns, [{c: r.get(c) for c in columns} for r in rows])
        )
    return sections


def _truthy(value) -> bool:
    if _is_missing(value):
        return False
    return bool(value)


def annotation_footnotes(fairness: "pd.DataFrame") -> Tuple[Dict[str, int], List[str]]:
    """Number the residents that carry notes: (marker map, footnote lines).

    The marker map (resident → 1-based number) lets the fairness table show a
    small reference instead of the full annotation text; the lines render as
    a Notes block after the tables ("1. Alice — [blackout …] [leave 2d comp]").
    """
    if "Notes" not in getattr(fairness, "columns", []):
        return {}, []
    markers: Dict[str, int] = {}
    lines: List[str] = []
    for record in fairness.to_dict("records"):
        note = record.get("Notes")
        if not _truthy(note):
            continue
        markers[record["Resident"]] = len(lines) + 1
        lines.append(f"{len(lines) + 1}. {record['Resident']} — {note}")
    return markers, lines


def report_header_lines(
    data: InputData,
    df,
    quality=None,
    validation_issues: Sequence[str] | None = None,
) -> List[str]:
    """Metadata lines for the report title block."""
    days = (data.end_date - data.start_date).days + 1
    lines = [
        f"Block {compact_date_range(data.start_date, data.end_date)} "
        f"({days} days) · {len(data.juniors)} juniors · {len(data.seniors)} seniors "
        f"· {len(data.shifts)} shift types",
        f"Generated {_date.today().strftime('%a %d %b %Y')}",
    ]
    attrs = getattr(df, "attrs", {}) or {}
    status = attrs.get("solver_status")
    detail = []
    if attrs.get("manually_edited"):
        detail.append("MANUALLY EDITED (not solver-certified)")
    elif status:
        wall = attrs.get("wall_time_sec")
        detail.append(f"Solver {status}" + (f" in {wall:.0f}s" if wall is not None else ""))
    if quality:
        detail.append(f"Quality {quality.get('score', 0)}/100")
        unfilled = quality.get("unfilled", 0)
        detail.append("all slots filled" if not unfilled else f"{unfilled} slot(s) unfilled")
    if validation_issues is not None:
        detail.append(
            "validation passed"
            if not validation_issues
            else f"{len(validation_issues)} validation issue(s)"
        )
    if detail:
        lines.append(" · ".join(detail))
    return lines


def legend_entries(color_mode: str, palette=None) -> List[Tuple[str | None, str]]:
    """(hex colour | None, label) pairs describing the schedule shading."""
    pal = {**DEFAULT_PALETTE, **(palette or {})}
    entries: List[Tuple[str | None, str]] = []
    if color_mode in ("auto", "weekend"):
        entries.append((pal["weekend"], "Weekend / holiday shift"))
    if color_mode in ("auto", "points"):
        entries.append((pal["points"], "Weekday shift (deeper = more points)"))
    if color_mode == "role":
        entries.append((pal["senior"], "Senior shift"))
        entries.append((pal["junior"], "Junior shift"))
    if color_mode == "role_weekend":
        entries.append((pal["senior"], "Senior shift (paler = weekday, darker = weekend)"))
        entries.append((pal["junior"], "Junior shift (paler than senior; darker = weekend)"))
    entries.append((pal["unfilled"], "Unfilled slot (no resident)"))
    entries.append((None, "Closed = shift stood down (not demand)"))
    return entries


def _authoritative_frame(display_df, authoritative_df=None):
    return authoritative_df if authoritative_df is not None else display_df


def _resolve_validation_issues(df, data: InputData, supplied=None) -> List[str]:
    if supplied is not None:
        return [str(issue) for issue in supplied]
    from .validation import validate_schedule

    return list(validate_schedule(df, data))


# --- Excel --------------------------------------------------------------------

def schedule_to_excel_bytes(
    df: "pd.DataFrame",
    data: InputData,
    points: Dict[str, ResidentPoints] | None = None,
    color_mode: str = "none",
    palette: Dict[str, str] | None = None,
    prior_ledger=None,
    *,
    authoritative_df=None,
    validation_issues: Sequence[str] | None = None,
    policy_snapshot: Mapping[str, object] | None = None,
    ledger_policy=None,
) -> bytes:
    """Serialise the schedule, fairness summary, and per-call audit to .xlsx.

    Sheet "Schedule" is the calendar grid (frozen header + date column, real
    date formatting, explicit "Unfilled" in empty slots, cells shaded to match
    the on-screen view); sheet "Fairness" is the per-resident summary with a
    wrapped Notes column; sheet "Per-call" (when the frame still carries its
    Date column) is the slot-by-slot audit. Requires ``openpyxl``.
    """
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    source_df = _authoritative_frame(df, authoritative_df)
    points = points if points is not None else calculate_points(source_df, data)
    fairness = build_fairness_frame(
        points, data, source_df, prior_ledger, ledger_policy=ledger_policy
    )
    issues = _resolve_validation_issues(source_df, data, validation_issues)
    policy = build_policy_snapshot_frame(
        data,
        source_df,
        issues,
        policy_snapshot,
        ledger_policy=ledger_policy,
        include_config_details=True,
        prior_ledger=prior_ledger,
    )

    # Render copy only: an empty shift cell prints as an explicit "Unfilled";
    # the caller's frame (used for fairness maths) is never touched.
    shift_labels = [s.label for s in data.shifts if s.label in df.columns]
    render_df = df.copy()
    render_columns = list(render_df.columns)
    for label in shift_labels:
        render_df[label] = [
            "Unfilled" if _is_missing(v) else v
            for v in render_df[label]
        ]
    render_df = spreadsheet_safe_frame(render_df)
    fairness = spreadsheet_safe_frame(fairness)
    policy = spreadsheet_safe_frame(policy)

    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    def _polish(worksheet, frame, wide_cols=(), wrap_cols=()):
        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.sheet_view.showGridLines = False
        for col_idx, column in enumerate(frame.columns, start=1):
            head = worksheet.cell(row=1, column=col_idx)
            head.fill = header_fill
            head.font = header_font
            letter = get_column_letter(col_idx)
            content_width = max(
                [len(_fmt(column))]
                + [len(_fmt(worksheet.cell(row=row_idx, column=col_idx).value))
                   for row_idx in range(2, len(frame) + 2)]
            ) + 2
            if column in wrap_cols:
                width = max(24, min(60, content_width))
                for row_idx in range(2, len(frame) + 2):
                    worksheet.cell(row=row_idx, column=col_idx).alignment = (
                        Alignment(wrap_text=True, vertical="top")
                    )
            elif column in wide_cols:
                width = max(16, min(36, content_width))
            else:
                width = max(10, min(28, content_width))
            worksheet.column_dimensions[letter].width = width

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        render_df.to_excel(writer, sheet_name="Schedule", index=False)
        fairness.to_excel(writer, sheet_name="Fairness", index=False)
        schedule_ws = writer.sheets["Schedule"]
        _polish(schedule_ws, render_df, wide_cols=("Date", "Day"))
        if "Date" in render_df.columns:
            date_col = list(render_df.columns).index("Date") + 1
            for row_idx in range(2, len(render_df) + 2):
                schedule_ws.cell(row=row_idx, column=date_col).number_format = "ddd dd mmm"
        _polish(
            writer.sheets["Fairness"], fairness,
            wide_cols=("Resident",), wrap_cols=("Notes",),
        )
        if color_mode and color_mode != "none":
            for (row_idx, label), hexcolor in schedule_cell_colors(source_df, data, color_mode, palette).items():
                if label in render_columns:
                    rgb = hexcolor.lstrip("#").upper()
                    schedule_ws.cell(
                        row=row_idx + 2,
                        column=render_columns.index(label) + 1,
                    ).fill = PatternFill(
                        start_color=rgb, end_color=rgb, fill_type="solid"
                    )
        if "Date" in source_df.columns:
            per_call = spreadsheet_safe_frame(build_assignment_frame(source_df, data))
            per_call.to_excel(writer, sheet_name="Per-call", index=False)
            _polish(
                writer.sheets["Per-call"], per_call,
                wide_cols=("Date", "Shift", "Status", "Resident"),
            )
        policy.to_excel(writer, sheet_name="Policy & validation", index=False)
        _polish(
            writer.sheets["Policy & validation"], policy,
            wide_cols=("Setting",), wrap_cols=("Value",),
        )
    return buffer.getvalue()


# --- PDF ------------------------------------------------------------------------

_WEEKEND_ROW_TINT = "#f6efdc"   # soft parchment behind weekend rows


def schedule_to_pdf_bytes(
    df: "pd.DataFrame",
    data: InputData,
    points: Dict[str, ResidentPoints] | None = None,
    color_mode: str = "none",
    palette: Dict[str, str] | None = None,
    prior_ledger=None,
    *,
    authoritative_df=None,
    validation_issues: Sequence[str] | None = None,
    policy_snapshot: Mapping[str, object] | None = None,
    ledger_policy=None,
) -> bytes:
    """Render the full report to a landscape-A4 PDF.

    Layout: title block (block dates, roster size, generated-on, solver
    status + quality) → colour legend → schedule grid (friendly dates,
    weekend rows tinted, explicit Unfilled, cells shaded to match the
    screen) → per-role fairness tables (curated print columns, footnote
    markers) → numbered Notes block. Column widths are content-aware (name
    columns wide, numerics narrow) instead of evenly split, and cell text is
    XML-escaped so names with ``&``/``<`` can't break the renderer.
    Requires ``reportlab``.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        KeepTogether,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    from .fairness import schedule_quality

    source_df = _authoritative_frame(df, authoritative_df)
    points = points if points is not None else calculate_points(source_df, data)
    fairness = build_fairness_frame(
        points, data, source_df, prior_ledger, ledger_policy=ledger_policy
    )
    quality = schedule_quality(source_df, data, points=points)
    issues = _resolve_validation_issues(source_df, data, validation_issues)
    policy = build_policy_snapshot_frame(
        data, source_df, issues, policy_snapshot, ledger_policy=ledger_policy
    )

    font_name, bold_font_name, unicode_font = _register_pdf_fonts()
    styles = getSampleStyleSheet()
    cell = ParagraphStyle("cell", fontName=font_name, fontSize=7, leading=8.5)
    cell_dim = ParagraphStyle(
        "cell_dim", parent=cell, fontName=font_name,
        textColor=colors.HexColor("#8a8378"),
    )
    head = ParagraphStyle(
        "head", fontName=bold_font_name, fontSize=7, leading=8.5,
        textColor=colors.white,
    )
    meta = ParagraphStyle(
        "meta", fontName=font_name, fontSize=8.5, leading=11,
        textColor=colors.HexColor("#4a4438"),
    )
    note_style = ParagraphStyle(
        "note", fontName=font_name, fontSize=7.5, leading=10
    )
    section_style = ParagraphStyle(
        "section", parent=styles["Heading2"], fontName=bold_font_name,
        keepWithNext=True,
    )

    page = landscape(A4)
    usable_width = page[0] - 2 * cm
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=page,
        leftMargin=cm,
        rightMargin=cm,
        topMargin=cm,
        bottomMargin=1.2 * cm,
        title="Idea Gold Schedule",
    )

    def _cell_par(value, marker=None):
        text = escape(_pdf_safe_text(value, unicode_font=unicode_font)) or "&nbsp;"
        if value in ("Unfilled", "Closed"):
            return Paragraph(text, cell_dim)
        if marker:
            text += f" <super>{marker}</super>"
        return Paragraph(text, cell)

    def _widths(columns, first_col_cm):
        n = len(columns) or 1
        if n == 1:
            return [usable_width]
        first = min(first_col_cm * cm, usable_width / 2)
        rest = (usable_width - first) / (n - 1)
        return [first] + [rest] * (n - 1)

    def _table(columns, rows, cell_bg=None, weekend_rows=None,
               first_col_cm=2.6, markers=None):
        header = [
            Paragraph(
                escape(_pdf_safe_text(c, unicode_font=unicode_font)) or "&nbsp;",
                head,
            )
            for c in columns
        ]
        body = []
        for row in rows:
            cells = []
            for col_idx, column in enumerate(columns):
                marker = None
                if markers and col_idx == 0:
                    marker = markers.get(row.get(column))
                cells.append(_cell_par(row.get(column), marker))
            body.append(cells)
        table = Table(
            [header] + body, colWidths=_widths(columns, first_col_cm), repeatRows=1
        )
        style = [
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#b9b2a4")),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]
        if weekend_rows:
            for row_idx in sorted(weekend_rows):
                style.append((
                    "BACKGROUND", (0, row_idx + 1), (-1, row_idx + 1),
                    colors.HexColor(_WEEKEND_ROW_TINT),
                ))
        elif not cell_bg:
            style.append((
                "ROWBACKGROUNDS", (0, 1), (-1, -1),
                [colors.white, colors.HexColor("#f2f0eb")],
            ))
        if cell_bg:
            # Per-cell shading to match the on-screen view; header stays dark.
            for (row_idx, label), hexcolor in cell_bg.items():
                if label in columns:
                    col = columns.index(label)
                    style.append(
                        ("BACKGROUND", (col, row_idx + 1), (col, row_idx + 1),
                         colors.HexColor(hexcolor))
                    )
        table.setStyle(TableStyle(style))
        return table

    def _legend_flowable():
        entries = legend_entries(color_mode, palette)
        item_width = usable_width / max(1, len(entries))
        cells = []
        for hexcolor, label in entries:
            item = Table(
                [[
                    "",
                    Paragraph(
                        escape(_pdf_safe_text(label, unicode_font=unicode_font)),
                        meta,
                    ),
                ]],
                colWidths=[0.22 * cm, max(0.5 * cm, item_width - 0.22 * cm)],
            )
            item_style = [
                ("LEFTPADDING", (0, 0), (-1, -1), 0),
                ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (0, 0), 0.4, colors.HexColor("#8a8378")),
            ]
            if hexcolor:
                item_style.append(
                    ("BACKGROUND", (0, 0), (0, 0), colors.HexColor(hexcolor))
                )
            item.setStyle(TableStyle(item_style))
            cells.append(item)
        table = Table([cells], colWidths=[item_width] * len(cells))
        table.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        return table

    def _footer(canvas, document):
        canvas.saveState()
        canvas.setFont(font_name, 7)
        canvas.setFillColor(colors.HexColor("#8a8378"))
        canvas.drawString(cm, 0.6 * cm, "Idea Gold Scheduler")
        canvas.drawRightString(page[0] - cm, 0.6 * cm, f"Page {document.page}")
        canvas.restoreState()

    schedule_bg = (
        schedule_cell_colors(source_df, data, color_mode, palette)
        if color_mode and color_mode != "none"
        else None
    )
    sched_cols, sched_rows, weekend_rows = schedule_print_view(df, data)
    markers, footnotes = annotation_footnotes(fairness)

    elements = [Paragraph("Idea Gold Schedule", styles["Title"])]
    for line in report_header_lines(data, source_df, quality, issues):
        elements.append(
            Paragraph(
                escape(_pdf_safe_text(line, unicode_font=unicode_font)), meta
            )
        )
    elements.append(Spacer(1, 6))
    elements.append(_legend_flowable())
    elements.append(Spacer(1, 6))
    elements.append(_table(
        sched_cols, sched_rows, cell_bg=schedule_bg, weekend_rows=weekend_rows,
    ))
    for title, columns, rows in fairness_print_sections(fairness, data):
        elements.append(Spacer(1, 12))
        elements.append(KeepTogether([
            Paragraph(
                "Fairness - "
                + escape(_pdf_safe_text(title, unicode_font=unicode_font)),
                section_style,
            ),
            Spacer(1, 3),
            _table(columns, rows, first_col_cm=3.4, markers=markers),
        ]))
    if footnotes:
        elements.append(Spacer(1, 10))
        elements.append(Paragraph("Notes", section_style))
        for line in footnotes:
            elements.append(
                Paragraph(
                    escape(_pdf_safe_text(line, unicode_font=unicode_font)),
                    note_style,
                )
            )
    elements.append(Spacer(1, 12))
    elements.append(KeepTogether([
        Paragraph("Policy &amp; validation", section_style),
        Spacer(1, 3),
        _table(list(policy.columns), policy.to_dict("records"), first_col_cm=4.2),
    ]))
    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()
