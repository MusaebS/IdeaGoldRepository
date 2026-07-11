"""Import of monthly availability requests from a form export (Excel/CSV).

Staff answer a monthly form with the periods they need to be free of calls;
the scheduler uploads the exported sheet and every valid row becomes a
*compensated* leave (full fair share kept — the request blocks the days but
is made up elsewhere, exactly like a hand-entered leave).

Expected columns (header names are matched case-insensitively and accept the
synonyms below): ``Name``, ``Start``, ``End``. One row per period; several
rows per person are fine; a missing/empty End means a single day. Dates may
be real date cells (typed .xlsx), ISO ``YYYY-MM-DD``, or ``DD/MM/YYYY``.

This module must stay importable without pandas/OR-Tools/Streamlit; the
``.xlsx`` reader/template guard their ``openpyxl`` import with a clear error.
"""
from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import List, Mapping, NamedTuple, Sequence

from .data_models import Leave
from .names import canonical_name

__all__ = [
    "AvailabilityRow",
    "TEMPLATE_HEADERS",
    "parse_availability_rows",
    "read_availability_csv",
    "read_availability_xlsx",
    "availability_template_csv",
    "availability_template_xlsx",
    "rows_to_leaves",
]

_NAME_HEADERS = {"name", "resident", "who"}
_START_HEADERS = {"start", "start date", "from", "first day"}
_END_HEADERS = {"end", "end date", "to", "until", "last day"}

TEMPLATE_HEADERS = ("Name", "Start", "End")
_TEMPLATE_EXAMPLE = ("Alice Example", "2026-08-05", "2026-08-07")


class AvailabilityRow(NamedTuple):
    """One parsed request row; ``error`` explains why an invalid row is skipped."""

    row_no: int
    raw_name: str
    name: str | None      # the matched roster name
    start: date | None
    end: date | None
    error: str | None


def _canon(text) -> str:
    """Use the shared Unicode-aware matching rule for headers and names."""
    return canonical_name(str(text))


def _find_column(fieldnames, synonyms) -> str | None:
    for field in fieldnames:
        if _canon(field) in synonyms:
            return field
    return None


def _parse_date(value) -> date | None:
    """Accept date/datetime cells, ISO strings, and DD/MM/YYYY variants."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value if value is not None else "").strip()
    if not text:
        return None
    try:
        return date.fromisoformat(text)
    except ValueError:
        pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(text)


def parse_availability_rows(
    rows: Sequence[Mapping], roster: Sequence[str]
) -> List[AvailabilityRow]:
    """Turn raw sheet rows into validated availability requests.

    Names match the roster after trimming/case-folding/whitespace-collapsing;
    every problem is reported per row (unknown name, unreadable or missing
    date, backwards range) so one bad answer never blocks the rest. Fully
    blank rows are skipped silently.
    """
    lookup = {_canon(name): name for name in roster}
    rows = list(rows)
    if not rows:
        return []
    fieldnames = list(rows[0].keys())
    name_col = _find_column(fieldnames, _NAME_HEADERS)
    start_col = _find_column(fieldnames, _START_HEADERS)
    end_col = _find_column(fieldnames, _END_HEADERS)
    if name_col is None or start_col is None:
        missing = [
            label
            for label, col in (("Name", name_col), ("Start", start_col))
            if col is None
        ]
        return [AvailabilityRow(
            0, "", None, None, None,
            f"Could not find the {' and '.join(missing)} column(s); expected "
            f"headers like {', '.join(TEMPLATE_HEADERS)}.",
        )]

    out: List[AvailabilityRow] = []
    for row_no, raw in enumerate(rows, start=2):  # row 1 is the header
        raw_name = str(raw.get(name_col) if raw.get(name_col) is not None else "").strip()
        raw_start = raw.get(start_col)
        raw_end = raw.get(end_col) if end_col is not None else None
        if not raw_name and _is_blank(raw_start) and _is_blank(raw_end):
            continue  # an empty spreadsheet line, not a mistake

        error = None
        name = lookup.get(_canon(raw_name)) if raw_name else None
        if not raw_name:
            error = "Missing name."
        elif name is None:
            error = f"'{raw_name}' is not on the roster."

        start = end = None
        if error is None:
            try:
                start = _parse_date(raw_start)
            except ValueError as exc:
                error = f"Unreadable start date '{exc}'."
            if error is None and start is None:
                error = "Missing start date."
        if error is None:
            try:
                end = _parse_date(raw_end)
            except ValueError as exc:
                error = f"Unreadable end date '{exc}'."
        if error is None:
            if end is None:
                end = start  # a single day
            elif start is not None and end < start:
                error = f"End {end} is before start {start}."

        out.append(AvailabilityRow(row_no, raw_name, name, start, end, error))
    return out


def _is_blank(value) -> bool:
    return value is None or not str(value).strip()


def read_availability_csv(text: str) -> List[Mapping]:
    """Rows of a CSV export as dicts (header row required)."""
    return [dict(row) for row in csv.DictReader(io.StringIO(text))]


def read_availability_xlsx(blob: bytes) -> List[Mapping]:
    """Rows of the first sheet of an .xlsx export as dicts (header row first)."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - exercised via RuntimeError path
        raise RuntimeError(
            "openpyxl is required to read .xlsx files; upload a CSV instead."
        ) from exc
    workbook = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    try:
        sheet = workbook.worksheets[0]
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            return []
        names = [str(h) if h is not None else "" for h in header]
        return [
            {names[i]: (values[i] if i < len(values) else None) for i in range(len(names))}
            for values in rows_iter
        ]
    finally:
        workbook.close()


def availability_template_csv() -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(TEMPLATE_HEADERS)
    writer.writerow(_TEMPLATE_EXAMPLE)
    return buffer.getvalue()


def availability_template_xlsx() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:  # pragma: no cover - exercised via RuntimeError path
        raise RuntimeError("openpyxl is required to build the .xlsx template.") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Availability"
    sheet.append(list(TEMPLATE_HEADERS))
    sheet.append(list(_TEMPLATE_EXAMPLE))
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def rows_to_leaves(parsed: Sequence[AvailabilityRow]) -> List[Leave]:
    """Valid rows become compensated leaves (the block is requested off but
    the fair share is kept, so the load is made up elsewhere or carried)."""
    return [
        Leave(row.name, row.start, row.end, True)
        for row in parsed
        if row.error is None and row.name and row.start and row.end
    ]
