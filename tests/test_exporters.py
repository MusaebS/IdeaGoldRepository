import sys, os
import io
from datetime import date

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import pytest

try:
    import pandas as pd
except Exception:  # pragma: no cover - fallback when pandas missing
    from model import optimiser as opt
    pd = opt.pd

from model.data_models import ShiftTemplate, InputData
from model.fairness import calculate_points
from model.exporters import (
    build_fairness_frame,
    schedule_to_excel_bytes,
    schedule_to_pdf_bytes,
)


def _sample():
    shifts = [
        ShiftTemplate(label="D", role="Junior", night_float=False, thu_weekend=False, points=1.0),
        ShiftTemplate(label="N", role="Senior", night_float=True, thu_weekend=False, points=2.0),
    ]
    data = InputData(
        start_date=date(2023, 1, 7),
        end_date=date(2023, 1, 8),
        shifts=shifts,
        juniors=["Alice"],
        seniors=["Bob"],
        nf_juniors=[],
        nf_seniors=["Bob"],
        leaves=[],
        rotators=[],
        min_gap=1,
    )
    df = pd.DataFrame([
        {"Date": date(2023, 1, 7), "Day": "Saturday", "D": "Alice", "N": "Bob"},
        {"Date": date(2023, 1, 8), "Day": "Sunday", "D": "Alice", "N": "Bob"},
    ])
    return df, data


def _df_and_data(**overrides):
    """Two-day, two-shift fixture; kwargs override the InputData fields.

    Passing ``seniors=["Bob"]`` moves Bob out of the junior list automatically
    so role-split tests don't have to restate the roster.
    """
    shifts = [
        ShiftTemplate(label="D", role="Junior", night_float=False, thu_weekend=False, points=1.0),
        ShiftTemplate(label="N", role="Junior", night_float=False, thu_weekend=False, points=2.0),
    ]
    base = dict(
        start_date=date(2023, 1, 7),
        end_date=date(2023, 1, 8),
        shifts=shifts,
        juniors=["Alice", "Bob"],
        seniors=[],
        nf_juniors=[],
        nf_seniors=[],
        leaves=[],
        rotators=[],
        min_gap=0,
    )
    base.update(overrides)
    if overrides.get("seniors") and "juniors" not in overrides:
        base["juniors"] = [p for p in base["juniors"] if p not in base["seniors"]]
    data = InputData(**base)
    df = pd.DataFrame([
        {"Date": date(2023, 1, 7), "Day": "Saturday", "D": "Alice", "N": "Bob"},
        {"Date": date(2023, 1, 8), "Day": "Sunday", "D": "Bob", "N": "Alice"},
    ])
    return df, data


def test_build_fairness_frame():
    df, data = _sample()
    points = calculate_points(df, data)
    frame = build_fairness_frame(points, data)
    by_name = {row["Resident"]: row for row in frame.to_dict("records")}
    assert by_name["Alice"]["Role"] == "Junior"
    assert by_name["Bob"]["Role"] == "Senior"
    assert by_name["Alice"]["Total"] == 2.0      # D 1.0 x 2 days
    assert by_name["Bob"]["Total"] == 4.0        # N 2.0 x 2 days (NF-eligible but uncovered = regular)
    assert by_name["Bob"]["NF duty (days)"] == 0  # no NF coverage configured
    assert "D" in frame.columns and "N" in frame.columns


def test_schedule_to_excel_bytes():
    pytest.importorskip("openpyxl")
    df, data = _sample()
    blob = schedule_to_excel_bytes(df, data)
    assert isinstance(blob, (bytes, bytearray))
    assert blob[:2] == b"PK"  # .xlsx is a zip archive


def test_schedule_to_pdf_bytes():
    pytest.importorskip("reportlab")
    df, data = _sample()
    blob = schedule_to_pdf_bytes(df, data)
    assert isinstance(blob, (bytes, bytearray))
    assert blob[:4] == b"%PDF"  # PDF magic number


def test_exports_accept_cosmetic_columns_and_palette():
    pytest.importorskip("openpyxl")
    pytest.importorskip("reportlab")
    df, data = _sample()
    # A purely cosmetic column (like an on-call team label) plus a custom palette
    # must not break the exports; fairness is computed from points, not columns.
    df = df.copy()
    df["On-call team"] = ["Red", "Blue"]
    points = calculate_points(df, data)
    frame = build_fairness_frame(points, data, df)
    assert "On-call team" not in frame.columns  # cosmetic column stays cosmetic
    xl = schedule_to_excel_bytes(
        df, data, points=points, color_mode="auto", palette={"weekend": "#123456"}
    )
    pdf = schedule_to_pdf_bytes(
        df, data, points=points, color_mode="role", palette={"senior": "#654321"}
    )
    assert xl[:2] == b"PK" and pdf[:4] == b"%PDF"


def test_fairness_frame_with_df_adds_counts_targets_and_notes():
    df, data = _sample()
    df.attrs["target_total_map"] = {"Alice": 3.0, "Bob": 3.0}
    df.attrs["target_weekend"] = {"Alice": 3.0, "Bob": 3.0}
    data.exempt_shifts = {"Alice": ["N"]}
    points = calculate_points(df, data)
    frame = build_fairness_frame(points, data, df)
    by_name = {row["Resident"]: row for row in frame.to_dict("records")}
    assert by_name["Alice"]["D n"] == 2 and by_name["Alice"]["N n"] == 0
    assert by_name["Bob"]["N n"] == 2
    assert by_name["Alice"]["Total target"] == 3.0
    assert by_name["Alice"]["Total dev"] == -1.0
    assert by_name["Alice"]["Weekend target"] == 3.0
    assert "NF target" not in by_name["Bob"]  # NF is no longer a balanced dimension
    assert "[exempt: N]" in by_name["Alice"]["Notes"]


def test_fairness_frame_with_prior_ledger_adds_cumulative():
    df, data = _sample()
    points = calculate_points(df, data)
    prior = {
        "Alice": {"total": 10.0, "weekend": 4.0, "night_float": 0.0,
                  "label_counts": {"D": 8}},
    }
    frame = build_fairness_frame(points, data, df, prior)
    by_name = {row["Resident"]: row for row in frame.to_dict("records")}
    assert by_name["Alice"]["Prior total"] == 10.0
    assert by_name["Alice"]["Cumulative total"] == 12.0
    assert by_name["Alice"]["Cumulative weekend"] == 6.0
    assert by_name["Alice"]["D n cum"] == 10  # 8 prior + 2 this block
    assert by_name["Bob"]["Prior total"] == 0.0
    assert by_name["Bob"]["Cumulative total"] == 4.0


def test_fairness_frame_without_new_args_keeps_old_shape():
    df, data = _sample()
    points = calculate_points(df, data)
    frame = build_fairness_frame(points, data)
    for absent in ("Total target", "Prior total", "Cumulative total", "D n", "Notes"):
        assert absent not in frame.columns


def test_build_assignment_frame_lists_every_slot():
    from model.exporters import build_assignment_frame

    _, data = _sample()
    # Built directly (not via .loc) so the pandas-stub CI job can run this too.
    df = pd.DataFrame([
        {"Date": date(2023, 1, 7), "Day": "Saturday", "D": "Alice", "N": "Bob"},
        {"Date": date(2023, 1, 8), "Day": "Sunday", "D": "Alice", "N": "Unfilled"},
    ])
    frame = build_assignment_frame(df, data)
    records = frame.to_dict("records")
    assert len(records) == 4  # 2 days x 2 shifts, unfilled included
    sat_n = next(r for r in records if r["Shift"] == "N" and r["Date"] == date(2023, 1, 7))
    assert sat_n["Resident"] == "Bob" and sat_n["Points"] == 2.0
    # NF eligibility is a shift property; without an active overlay this is a
    # regular call that earns points, not night-float duty.
    assert sat_n["Weekend"] is True and sat_n["Night float"] is False
    assert sat_n["NF-eligible shift"] is True
    assert sat_n["Status"] == "Regular assignment"
    sun_n = next(r for r in records if r["Shift"] == "N" and r["Date"] == date(2023, 1, 8))
    assert sun_n["Resident"] == "Unfilled" and sun_n["Points"] == 0.0


# --- print views / report helpers ----------------------------------------------

def test_schedule_print_view_merges_dates_and_labels_unfilled():
    from model.exporters import schedule_print_view

    df, data = _df_and_data()
    records = df.to_dict("records")
    records[1]["D"] = None  # a genuine gap
    columns, rows, weekend_rows = schedule_print_view(pd.DataFrame(records), data)
    assert columns[0] == "Date" and "Day" not in columns
    assert rows[0]["Date"].startswith(("Sat", "Sun", "Mon", "Tue", "Wed", "Thu", "Fri"))
    assert rows[1]["D"] == "Unfilled"  # explicit, not a blank cell


def test_schedule_print_view_flags_weekend_rows():
    from model.exporters import schedule_print_view

    shifts = [ShiftTemplate(label="D", role="Junior", night_float=False, thu_weekend=False, points=1.0)]
    data = InputData(
        start_date=date(2023, 1, 6), end_date=date(2023, 1, 8),  # Fri..Sun
        shifts=shifts, juniors=["A", "B"], seniors=[], nf_juniors=[], nf_seniors=[],
        leaves=[], rotators=[], min_gap=0,
    )
    df = pd.DataFrame([
        {"Date": date(2023, 1, 6), "Day": "Friday", "D": "A"},
        {"Date": date(2023, 1, 7), "Day": "Saturday", "D": "B"},
        {"Date": date(2023, 1, 8), "Day": "Sunday", "D": "A"},
    ])
    _, _, weekend_rows = schedule_print_view(df, data)
    assert weekend_rows == {1, 2}


def test_fairness_print_sections_split_by_role_and_curated():
    from model.exporters import fairness_print_sections

    df, data = _df_and_data(seniors=["Bob"])
    points = calculate_points(df, data)
    fairness = build_fairness_frame(points, data, df)
    sections = fairness_print_sections(fairness, data)
    titles = [t for t, _, _ in sections]
    assert titles == ["Juniors", "Seniors"]
    junior_cols = sections[0][1]
    assert "Notes" not in junior_cols and "Role" not in junior_cols
    assert "Resident" in junior_cols and "Total" in junior_cols
    # Per-label COUNT columns print; raw per-label point columns stay in CSV.
    assert "D n" in junior_cols and "D" not in junior_cols
    # NF duty is dropped when nobody has any.
    assert "NF duty (days)" not in junior_cols


def test_annotation_footnotes_number_only_noted_residents():
    from model.exporters import annotation_footnotes

    df, data = _df_and_data(exempt_shifts={"Alice": ["N"]})
    points = calculate_points(df, data)
    fairness = build_fairness_frame(points, data, df)
    markers, lines = annotation_footnotes(fairness)
    assert markers == {"Alice": 1}
    alice_notes = next(
        r["Notes"] for r in fairness.to_dict("records") if r["Resident"] == "Alice"
    )
    assert lines == [f"1. Alice — {alice_notes}"]


def test_report_header_and_legend():
    from model.exporters import legend_entries, report_header_lines

    df, data = _df_and_data()
    df.attrs["solver_status"] = "OPTIMAL"
    df.attrs["wall_time_sec"] = 3.0
    lines = report_header_lines(data, df, {"score": 100.0, "unfilled": 0})
    text = " ".join(lines)
    assert "juniors" in text and "Generated" in text
    assert "Solver OPTIMAL in 3s" in text and "Quality 100.0/100" in text
    legend = legend_entries("auto")
    labels = [label for _, label in legend]
    assert any("Unfilled" in label for label in labels)
    assert any("Weekend" in label for label in labels)


def test_build_cumulative_frame_segments():
    from model.exporters import build_cumulative_frame

    df, data = _df_and_data()
    points = calculate_points(df, data)
    prior = {"Alice": {"total": 5.0, "weekend": 2.0}}
    rows = build_cumulative_frame(points, prior, data).to_dict("records")
    alice = [r for r in rows if r["Resident"] == "Alice"]
    assert {r["Segment"] for r in alice} == {"Prior blocks", "This block"}
    prior_row = next(r for r in alice if r["Segment"] == "Prior blocks")
    assert prior_row["Points"] == 5.0
    assert prior_row["Cumulative"] == 5.0 + points["Alice"]["total"]
    # Bob has no history: prior segment is zero, not missing.
    bob_prior = [
        r for r in rows
        if r["Resident"] == "Bob" and r["Segment"] == "Prior blocks"
    ]
    assert len(bob_prior) == 1 and bob_prior[0]["Points"] == 0.0


def _with_gap(df):
    """A frame whose first D cell is a genuine solver-style None gap."""
    records = df.to_dict("records")
    records[0]["D"] = None
    return pd.DataFrame(records)


def test_pdf_with_notes_ledger_and_gaps_still_renders():
    pytest.importorskip("reportlab")
    from model.data_models import Blackout

    df, data = _df_and_data(
        seniors=["Bob"],
        blackouts=[Blackout(None, ("Alice",), date(2023, 1, 7), date(2023, 1, 8))],
    )
    prior = {"Alice": {"total": 5.0, "weekend": 2.0}}
    out = schedule_to_pdf_bytes(_with_gap(df), data, color_mode="auto", prior_ledger=prior)
    assert out[:4] == b"%PDF"


def test_excel_gains_per_call_sheet_and_unfilled_labels():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    df, data = _df_and_data()
    df = _with_gap(df)
    out = schedule_to_excel_bytes(df, data)
    book = load_workbook(io.BytesIO(out))
    assert set(book.sheetnames) >= {"Schedule", "Fairness", "Per-call"}
    schedule = book["Schedule"]
    header = [c.value for c in schedule[1]]
    values = [c.value for c in schedule[2]]
    assert values[header.index("D")] == "Unfilled"
    assert schedule.freeze_panes == "B2"


def test_per_call_status_and_points_match_regular_fairness_accounting():
    from model.exporters import build_assignment_frame

    shift = ShiftTemplate(
        label="N", role="Senior", night_float=True, thu_weekend=False, points=2.0
    )
    data = InputData(
        start_date=date(2023, 1, 2), end_date=date(2023, 1, 5), shifts=[shift],
        juniors=[], seniors=["Bob"], nf_juniors=[], nf_seniors=["Bob"],
        leaves=[], rotators=[], min_gap=0,
    )
    df = pd.DataFrame([
        {"Date": date(2023, 1, 2), "N": "Bob"},
        {"Date": date(2023, 1, 3), "N": "Bob"},
        {"Date": date(2023, 1, 4), "N": "Closed"},
        {"Date": date(2023, 1, 5), "N": "Unfilled"},
    ])
    df.attrs["nf_cells"] = {"2023-01-02": {"N": "Bob"}}
    df.attrs["closed_cells"] = {"2023-01-04": ["N"]}
    rows = build_assignment_frame(df, data).to_dict("records")
    by_day = {row["Date"]: row for row in rows}
    assert by_day[date(2023, 1, 2)]["Status"] == "Night float overlay"
    assert by_day[date(2023, 1, 2)]["Points"] == 0.0
    assert by_day[date(2023, 1, 2)]["Night float"] is True
    assert by_day[date(2023, 1, 3)]["Status"] == "Regular assignment"
    assert by_day[date(2023, 1, 3)]["Points"] == 2.0
    assert by_day[date(2023, 1, 3)]["Night float"] is False
    assert by_day[date(2023, 1, 4)]["Status"] == "Closed"
    assert by_day[date(2023, 1, 4)]["Points"] == 0.0
    assert by_day[date(2023, 1, 5)]["Status"] == "Unfilled"
    assert by_day[date(2023, 1, 5)]["Points"] == 0.0
    assert all(row["Nominal points"] == 2.0 for row in rows)


def test_fairness_includes_configured_unassigned_labels_and_attr_targets():
    df, data = _df_and_data()
    data.shifts.append(ShiftTemplate(
        label="Unused", role="Junior", night_float=False,
        thu_weekend=False, points=3.0,
    ))
    df.attrs["target_label"] = {("Alice", "Unused"): 2.0}
    frame = build_fairness_frame(calculate_points(df, data), data, df)
    alice = next(row for row in frame.to_dict("records") if row["Resident"] == "Alice")
    assert alice["Unused"] == 0.0 and alice["Unused n"] == 0
    assert alice["Unused target"] == 2.0 and alice["Unused dev"] == -2.0


def test_role_sections_omit_opposite_role_shift_columns():
    from model.exporters import fairness_print_sections

    shifts = [
        ShiftTemplate("JD", "Junior", False, False, 1.0),
        ShiftTemplate("SD", "Senior", False, False, 1.0),
    ]
    data = InputData(
        start_date=date(2023, 1, 2), end_date=date(2023, 1, 2), shifts=shifts,
        juniors=["J"], seniors=["S"], nf_juniors=[], nf_seniors=[],
        leaves=[], rotators=[], min_gap=0,
    )
    df = pd.DataFrame([{"Date": date(2023, 1, 2), "JD": "J", "SD": "S"}])
    sections = fairness_print_sections(
        build_fairness_frame(calculate_points(df, data), data, df), data
    )
    by_title = {title: columns for title, columns, _rows in sections}
    assert "JD n" in by_title["Juniors"] and "SD n" not in by_title["Juniors"]
    assert "SD n" in by_title["Seniors"] and "JD n" not in by_title["Seniors"]


def test_authoritative_frame_drives_hidden_shift_export_accounting():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    source, data = _df_and_data()
    source.attrs["target_total_map"] = {"Alice": 3.0, "Bob": 3.0}
    display = source[["Day", "D"]].copy()  # Date and N hidden cosmetically
    out = schedule_to_excel_bytes(
        display, data, authoritative_df=source, validation_issues=[]
    )
    book = load_workbook(io.BytesIO(out), data_only=False)
    assert "N" not in [cell.value for cell in book["Schedule"][1]]
    per_call = list(book["Per-call"].values)
    headers = list(per_call[0])
    n_rows = [row for row in per_call[1:] if row[headers.index("Shift")] == "N"]
    assert {row[headers.index("Resident")] for row in n_rows} == {"Alice", "Bob"}
    fairness = list(book["Fairness"].values)
    fair_headers = list(fairness[0])
    alice = next(row for row in fairness[1:] if row[0] == "Alice")
    assert alice[fair_headers.index("N")] == 2.0


def test_spreadsheet_formula_text_is_neutralised_in_helpers_and_excel():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook
    from model.exporters import spreadsheet_safe_text

    for dangerous in ("=2+2", "+cmd", "-cmd", "@SUM(A1:A2)", "\tcmd", "\rcmd"):
        assert spreadsheet_safe_text(dangerous).startswith("'")
    assert spreadsheet_safe_text(2.0) == 2.0
    shift = ShiftTemplate("D", "Junior", False, False, 1.0)
    data = InputData(
        start_date=date(2023, 1, 2), end_date=date(2023, 1, 2), shifts=[shift],
        juniors=["=2+2"], seniors=[], nf_juniors=[], nf_seniors=[],
        leaves=[], rotators=[], min_gap=0,
    )
    df = pd.DataFrame([{"Date": date(2023, 1, 2), "D": "=2+2"}])
    out = schedule_to_excel_bytes(df, data, validation_issues=[])
    book = load_workbook(io.BytesIO(out), data_only=False)
    resident = book["Schedule"]["B2"]
    assert resident.data_type == "s" and resident.value == "'=2+2"

    dangerous_shift = ShiftTemplate("=2+2", "Junior", False, False, 1.0)
    header_data = InputData(
        start_date=date(2023, 1, 2), end_date=date(2023, 1, 2),
        shifts=[dangerous_shift], juniors=["A"], seniors=[],
        nf_juniors=[], nf_seniors=[], leaves=[], rotators=[], min_gap=0,
    )
    header_df = pd.DataFrame([{"Date": date(2023, 1, 2), "=2+2": "A"}])
    header_book = load_workbook(
        io.BytesIO(
            schedule_to_excel_bytes(header_df, header_data, validation_issues=[])
        ),
        data_only=False,
    )
    header = header_book["Schedule"]["B1"]
    assert header.data_type == "s" and header.value == "'=2+2"


def test_excel_widths_follow_long_resident_content():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    long_name = "Resident With A Meaningfully Long Display Name"
    shift = ShiftTemplate("D", "Junior", False, False, 1.0)
    data = InputData(
        start_date=date(2023, 1, 2), end_date=date(2023, 1, 2), shifts=[shift],
        juniors=[long_name], seniors=[], nf_juniors=[], nf_seniors=[],
        leaves=[], rotators=[], min_gap=0,
    )
    df = pd.DataFrame([{"Date": date(2023, 1, 2), "D": long_name}])
    book = load_workbook(io.BytesIO(schedule_to_excel_bytes(df, data, validation_issues=[])))
    fair = book["Fairness"]
    resident_col = next(cell.column_letter for cell in fair[1] if cell.value == "Resident")
    assert fair.column_dimensions[resident_col].width >= 30


def test_cumulative_frame_distinguishes_actual_and_policy_adjusted_standing():
    from model.exporters import build_cumulative_frame

    df, data = _df_and_data(extra_points={"Alice": 2.0})
    points = calculate_points(df, data)
    rows = build_cumulative_frame(
        points, {"Alice": {"total": 5.0, "weekend": 0.0}}, data
    ).to_dict("records")
    alice = next(row for row in rows if row["Resident"] == "Alice")
    assert alice["Cumulative"] == 5.0 + points["Alice"]["total"]
    assert alice["Policy-adjusted cumulative"] == alice["Cumulative"] - 2.0
    assert "Actual points" in alice["Standing basis"]


def test_first_block_policy_adjustment_is_visible_without_prior_ledger():
    df, data = _df_and_data(extra_points={"Alice": 2.0})
    frame = build_fairness_frame(calculate_points(df, data), data, df)
    # Iterate records (not .set_index/.loc) so the no-pandas stub CI job runs it.
    alice = next(r for r in frame.to_dict("records") if r["Resident"] == "Alice")
    assert "Cumulative total" in frame.columns
    assert "Policy-adjusted cumulative total" in frame.columns
    assert alice["Policy-adjusted cumulative total"] == pytest.approx(
        alice["Cumulative total"] - 2.0
    )


def test_excel_policy_sheet_contains_exact_config_rows():
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    df, data = _df_and_data(extra_points={"Alice": 2.0})
    df.attrs["label_carryover"] = False
    out = schedule_to_excel_bytes(df, data, validation_issues=[])
    book = load_workbook(io.BytesIO(out), data_only=False)
    rows = {
        setting: value
        for setting, value in book["Policy & validation"].iter_rows(
            min_row=2, values_only=True
        )
    }
    assert rows["Per-shift-type ledger carryover"] == "Disabled"
    assert '"Alice":2.0' in rows["Config JSON: extra_points"]
    assert "Config JSON: shifts" in rows


def test_report_header_marks_manual_edits_as_not_solver_certified():
    from model.exporters import report_header_lines

    df, data = _df_and_data()
    df.attrs["solver_status"] = "OPTIMAL"
    df.attrs["manually_edited"] = True
    text = " ".join(report_header_lines(data, df, validation_issues=["bad edit"]))
    assert "MANUALLY EDITED" in text and "Solver OPTIMAL" not in text
    assert "1 validation issue" in text


def test_pdf_safe_fallback_and_nan_omission():
    from model.exporters import _fmt, _pdf_safe_text

    # These helpers are dependency-free and must work in the stub CI job.
    assert _fmt(float("nan")) == ""
    rendered = _pdf_safe_text("محمد")
    assert "محمد" not in rendered and "U+0645" in rendered


def test_register_pdf_fonts_unicode_shaping():
    # Font registration imports reportlab; only exercise it where installed.
    pytest.importorskip("reportlab")
    from model.exporters import _pdf_safe_text, _register_pdf_fonts

    _normal, _bold, unicode_supported = _register_pdf_fonts()
    if unicode_supported:
        assert "U+" not in _pdf_safe_text("محمد", unicode_font=True)
